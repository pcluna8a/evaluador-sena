import streamlit as st
import google.generativeai as genai
import PyPDF2
import pandas as pd
import json
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestor de Idoneidad SENA 2026", page_icon="🇨🇴", layout="wide")

# --- GESTIÓN DE SEGURIDAD (API KEY) ---
try:
    api_key = st.secrets["GOOGLE_API_KEY"]
except:
    # Si no está en secretos, la pedimos manual (útil para pruebas locales)
    with st.sidebar:
        st.warning("Configuración Local detectada")
        api_key = st.text_input("Tu Google API Key:", type="password")

# --- FUNCIONES CENTRALES ---

def extraer_texto_pdf(archivo):
    """Extrae el texto crudo del PDF para que la IA lo lea."""
    try:
        pdf_reader = PyPDF2.PdfReader(archivo)
        texto = ""
        for pagina in pdf_reader.pages:
            txt = pagina.extract_text()
            if txt: texto += txt + "\n"
        return texto
    except Exception as e:
        return "Error lectura PDF"

def consultar_cerebro_ia(texto_cv, requisitos):
    """
    El núcleo inteligente. Evalúa alternativas, fechas antiguas y genera veredicto.
    """
    if not api_key: return None
    
    genai.configure(api_key=api_key)
    # Usamos flash para velocidad, o pro para mayor razonamiento si tienes acceso
    model = genai.GenerativeModel('gemini-1.5-flash')
    
    prompt = f"""
    Actúa como Coordinador de Talento Humano del SENA. Tu tarea es diligenciar el formato de idoneidad.
    
    REGLAS DE NEGOCIO CRÍTICAS:
    1. **Fecha de Grado:** Es el punto de partida. Solo cuenta la experiencia POSTERIOR a esta fecha.
    2. **Fechas Históricas:** Acepta fechas antiguas (ej: 1989, 1995, 2005). El sistema anterior fallaba con esto, tú debes sumarlas correctamente.
    3. **Alternativas:** El perfil puede tener "Alternativa 1" (ej: Título + Esp) o "Alternativa 2" (ej: Título + Exp). Si cumple CUALQUIERA, el veredicto es "CUMPLE".
    4. **Sumatoria:** Suma los meses de experiencia válida de todas las certificaciones detectadas post-grado.
    
    PERFIL REQUERIDO:
    {requisitos}
    
    HOJA DE VIDA DEL CANDIDATO:
    {texto_cv}
    
    SALIDA JSON OBLIGATORIA (Sin markdown):
    {{
        "nombre": "Nombre completo normalizado",
        "cedula": "Número de documento sin puntos",
        "fecha_grado": "DD/MM/AAAA",
        "veredicto": "CUMPLE" o "NO CUMPLE",
        "meses_exp": (Número entero de meses válidos),
        "alternativa": "Indica 'Alternativa 1', 'Alternativa 2' o 'N/A'",
        "empresas": "Lista resumida de empresas válidas",
        "observacion": "Breve justificación técnica del concepto (Máx 20 palabras)"
    }}
    """
    
    try:
        response = model.generate_content(prompt)
        clean_json = response.text.replace("```json", "").replace("```", "").strip()
        return json.loads(clean_json)
    except Exception as e:
        return {"nombre": "Error AI", "observacion": str(e)}

def llenar_plantilla_excel(dataframe, archivo_plantilla):
    """
    Toma la plantilla 2026_IDONEIDAD.xltx y vacía los datos en las celdas.
    """
    # Cargamos la plantilla en memoria
    wb = openpyxl.load_workbook(archivo_plantilla)
    ws = wb.active # Toma la primera hoja activa
    
    # Encontramos la primera fila vacía (Asumiendo que hay encabezados)
    # Generalmente los formatos SENA tienen encabezados en las primeras 5-7 filas.
    # Empezaremos a buscar espacio desde la fila 5 en adelante.
    fila_inicial = 1
    for row in range(1, 20):
        if ws.cell(row=row, column=1).value is None and ws.cell(row=row+1, column=1).value is None:
             # Si encontramos 2 filas vacías seguidas, asumimos que ahí empieza la data
             # O simplemente, buscamos la ultima fila llena + 1
             fila_inicial = ws.max_row + 1
             break
    
    start_row = ws.max_row + 1 if ws.max_row > 1 else 2
    
    # Estilos básicos (Bordes delgados para que se vea bien)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Iteramos sobre los datos de la IA y los escribimos
    # AJUSTE: Mapea estas columnas al ORDEN de tu Excel.
    # Asumo este orden estándar: [Nombre, Cedula, Fecha Grado, Meses, Empresas, Alternativa, Concepto, Observacion]
    
    for index, row in dataframe.iterrows():
        # Escribimos celda por celda
        ws.cell(row=start_row, column=1, value=row['nombre']).border = thin_border
        ws.cell(row=start_row, column=2, value=row['cedula']).border = thin_border
        ws.cell(row=start_row, column=3, value=row['fecha_grado']).border = thin_border
        ws.cell(row=start_row, column=4, value=row['meses_exp']).border = thin_border
        ws.cell(row=start_row, column=5, value=row['empresas']).border = thin_border
        ws.cell(row=start_row, column=6, value=row['alternativa']).border = thin_border
        
        # Celda de Concepto (Colorizada)
        celda_concepto = ws.cell(row=start_row, column=7, value=row['veredicto'])
        celda_concepto.border = thin_border
        celda_concepto.alignment = Alignment(horizontal='center')
        
        # Colores básicos (No usamos estilos complejos para asegurar compatibilidad)
        # Verde si CUMPLE, Rojo/Naranja si NO
        # Nota: Openpyxl requiere códigos HEX ARGB
        
        ws.cell(row=start_row, column=8, value=row['observacion']).border = thin_border
        
        start_row += 1

    # Guardamos en un buffer virtual para descargar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- INTERFAZ GRÁFICA ---

st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/8/83/Sena_Colombia_logo.svg/1200px-Sena_Colombia_logo.svg.png", width=120)
st.title("Sistema Experto de Evaluación - Idoneidad 2026")
st.markdown("Plataforma AI para validación de perfiles, cálculo de experiencia histórica y generación de informes.")

col_izq, col_der = st.columns([1, 1])

with col_izq:
    st.subheader("1. Configuración del Perfil (Norma)")
    st.info("Pegue aquí los requisitos completos, incluyendo Alternativa 1 y 2.")
    requisitos_txt = st.text_area("Requisitos:", height=200, placeholder="Ej: Título Profesional + 24 meses... O Alternativa 2...")

with col_der:
    st.subheader("2. Plantilla Institucional")
    st.info("Cargue el archivo '2026_IDONEIDAD.xltx' o '.xlsx'")
    archivo_plantilla = st.file_uploader("Formato Excel Base", type=["xlsx", "xltx"])

st.markdown("---")
st.subheader("3. Lote de Hojas de Vida")
archivos_pdf = st.file_uploader("Seleccione las HVs a evaluar (PDF)", type="pdf", accept_multiple_files=True)

# --- BOTÓN PRINCIPAL ---

if st.button("🚀 EJECUTAR EVALUACIÓN Y LLENAR FORMATO"):
    if not api_key or not requisitos_txt or not archivos_pdf:
        st.error("⚠️ Faltan datos: Asegúrese de tener API Key, Requisitos y Archivos cargados.")
    else:
        # Contenedores para resultados
        resultados = []
        barra = st.progress(0)
        status = st.empty()
        total = len(archivos_pdf)
        
        for i, pdf in enumerate(archivos_pdf):
            status.text(f"Analizando candidato {i+1}/{total}: {pdf.name}...")
            
            # 1. Leer
            texto = extraer_texto_pdf(pdf)
            
            # 2. Pensar (AI)
            datos = consultar_cerebro_ia(texto, requisitos_txt)
            
            if datos:
                datos['archivo_origen'] = pdf.name
                resultados.append(datos)
            
            barra.progress((i + 1) / total)
            
        status.success("✅ Análisis finalizado. Consolidando archivo Excel...")
        
        # CREAMOS EL DATAFRAME
        if resultados:
            df = pd.DataFrame(resultados)
            
            # Mostramos un adelanto en pantalla
            st.write("### Vista Previa de Resultados")
            st.dataframe(df)
            
            # GENERAMOS EL EXCEL
            if archivo_plantilla:
                # Si el usuario subió plantilla, usamos la lógica de inyección
                excel_final = llenar_plantilla_excel(df, archivo_plantilla)
                nombre_archivo = "2026_IDONEIDAD_DILIGENCIADO.xlsx"
            else:
                # Si no, generamos uno genérico
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)
                excel_final = output
                nombre_archivo = "Reporte_General_Idoneidad.xlsx"
            
            # BOTÓN DE DESCARGA
            st.download_button(
                label="📥 DESCARGAR ARCHIVO CONSOLIDADO",
                data=excel_final,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # RESUMEN VISUAL
            c1, c2 = st.columns(2)
            cumplen = df[df['veredicto'] == 'CUMPLE'].shape[0]
            no_cumplen = df[df['veredicto'] == 'NO CUMPLE'].shape[0]
            
            c1.metric("Candidatos QUE CUMPLEN", cumplen)
            c2.metric("NO CUMPLEN / REVISAR", no_cumplen)
            
        else:
            st.error("No se pudo extraer información de los documentos.")
