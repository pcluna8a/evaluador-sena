
import streamlit as st
import google.generativeai as genai
from pypdf import PdfReader
from PIL import Image
import os
import json
import openpyxl
from openpyxl.styles import Alignment
import io
import re
import pandas as pd

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(
    page_title="Evaluador SENA 2025",
    page_icon="✅",
    layout="wide"
)

# --- SIDEBAR / CONFIGURACIÓN ---
with st.sidebar:
    # Logo Oficial SENA
    st.image("https://www.sena.edu.co/Style%20Library/alayout/images/logoSena.png", width=180)
    st.markdown("### Configuración")
    
    # Dark Mode Toggle
    dark_mode = st.toggle("🌙 Modo Oscuro", value=False)
    
    # API Key Management
    api_key = os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        api_key = st.text_input("🔑 Google API Key", type="password")
    
    if api_key:
        genai.configure(api_key=api_key)
    else:
        st.warning("⚠️ API Key requerida.")

    # --- SECCIÓN COMPARTIR ---
    st.markdown("---")
    st.markdown("### 🔗 Compartir")
    import pyshorteners
    
    app_url = st.text_input("URL de la App:", placeholder="https://...")
    if app_url and st.button("Generar Link Corto"):
        try:
            s = pyshorteners.Shortener()
            st.code(s.tinyurl.short(app_url), language="text")
        except:
            st.error("Error al generar link.")

# --- ESTILOS CSS DINÁMICOS (MODERNO & INSTITUCIONAL) ---
if dark_mode:
    bg_color = "#121212"
    text_color = "#E0E0E0"
    card_bg = "#1E1E1E"
    input_bg = "#2C2C2C"
    border_color = "#444"
    header_text = "#FFFFFF"
else:
    bg_color = "#F4F6F8"
    text_color = "#2C3E50"
    card_bg = "#FFFFFF"
    input_bg = "#FFFFFF"
    border_color = "#E0E0E0"
    header_text = "#00324D"

st.markdown(f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');

        :root {{
            --sena-green: #39A900;
            --sena-dark-blue: #00324D;
            --sena-light-blue: #82DEF0;
            --sena-yellow: #FFCE40;
            --sena-dark-grey: #385C57;
            
            --bg-color: {bg_color};
            --text-color: {text_color};
            --card-bg: {card_bg};
            --input-bg: {input_bg};
            --border-color: {border_color};
            --header-text: {header_text};
        }}

        .stApp {{
            background-color: var(--bg-color);
            font-family: 'Inter', sans-serif;
            color: var(--text-color);
        }}
        
        #MainMenu, footer, header {{visibility: hidden;}}

        /* Encabezado Moderno */
        .main-header {{
            background: linear-gradient(135deg, var(--sena-green) 0%, #2E8B00 100%);
            padding: 2rem;
            text-align: center;
            color: white;
            border-bottom: 6px solid var(--sena-dark-blue);
            margin-bottom: 2.5rem;
            box-shadow: 0 10px 20px rgba(57, 169, 0, 0.15);
            border-radius: 0 0 20px 20px;
        }}
        .main-header h1 {{ 
            color: white; margin: 0; font-size: 2.2rem; font-weight: 700; letter-spacing: -0.5px;
        }}
        .main-header h2 {{ 
            color: rgba(255,255,255,0.9); margin-top: 0.5rem; font-size: 1.1rem; font-weight: 400;
        }}

        /* Tarjetas (Cards) */
        div[data-testid="stVerticalBlock"] > div {{
            background-color: var(--card-bg);
            padding: 1.5rem;
            border-radius: 16px;
            border: 1px solid var(--border-color);
            box-shadow: 0 4px 12px rgba(0,0,0,0.03);
            transition: transform 0.2s ease;
        }}
        
        h3 {{
            color: var(--header-text) !important;
            font-weight: 700 !important;
            font-size: 1.1rem !important;
            border-bottom: 2px solid var(--sena-yellow);
            padding-bottom: 0.8rem;
            margin-bottom: 1.5rem;
            display: inline-block;
        }}

        /* Inputs Modernos */
        .stTextInput input, .stTextArea textarea {{
            background-color: var(--input-bg);
            color: var(--text-color);
            border: 1px solid var(--border-color);
            border-radius: 10px;
            padding: 14px;
        }}
        .stTextInput input:focus, .stTextArea textarea:focus {{
            border-color: var(--sena-green) !important;
            box-shadow: 0 0 0 2px rgba(57, 169, 0, 0.1) !important;
        }}

        /* Botón Principal */
        .stButton button {{
            background: var(--sena-dark-blue) !important;
            color: white !important;
            border: none !important;
            padding: 16px 32px !important;
            font-size: 1.1rem !important;
            font-weight: 600 !important;
            border-radius: 12px !important;
            width: 100% !important;
            box-shadow: 0 4px 12px rgba(0, 50, 77, 0.2) !important;
            transition: all 0.3s ease !important;
        }}
        .stButton button:hover {{
            background: #004466 !important;
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 16px rgba(0, 50, 77, 0.3) !important;
        }}

        /* Resultados */
        .result-container {{
            background-color: var(--card-bg);
            padding: 2rem;
            border-radius: 16px;
            border-left: 6px solid var(--sena-green);
            box-shadow: 0 8px 24px rgba(0,0,0,0.06);
            margin-top: 2rem;
        }}
        
        .result-container h1, .result-container h2 {{ color: var(--header-text); border: none; }}
        
        /* Tablas */
        .result-container table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            margin: 20px 0;
            border-radius: 10px;
            overflow: hidden;
            border: 1px solid var(--border-color);
        }}
        .result-container th {{
            background-color: var(--sena-dark-blue);
            color: white;
            font-weight: 600;
            padding: 14px;
            text-align: left;
        }}
        .result-container td {{
            padding: 14px;
            border-bottom: 1px solid var(--border-color);
            color: var(--text-color);
        }}
        .result-container tr:last-child td {{ border-bottom: none; }}
    </style>
""", unsafe_allow_html=True)

# --- HEADER ---
st.markdown("""
    <div class="main-header">
        <h1>Evaluador de Idoneidad SENA 2025</h1>
        <h2>Análisis Inteligente de Perfiles y Experiencia</h2>
    </div>
""", unsafe_allow_html=True)

# --- LÓGICA DE NEGOCIO ---
def extraer_texto_pdf(uploaded_file):
    try:
        reader = PdfReader(uploaded_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"[Error PDF: {str(e)}]"

def cargar_imagen(uploaded_file):
    try:
        return Image.open(uploaded_file)
    except Exception as e:
        return None

def optimize_image(image, max_size=(1024, 1024)):
    """Redimensiona y optimiza la imagen para reducir latencia."""
    try:
        # Redimensionar si es muy grande (mantiene relación de aspecto)
        if image.width > max_size[0] or image.height > max_size[1]:
            image.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # Convertir a RGB (elimina canal Alpha que pesa más)
        if image.mode != 'RGB':
            image = image.convert('RGB')
            
        return image
    except Exception as e:
        return None

def fill_excel_template(data_json, template_path="2026_IDONEIDAD_NEW.xlsx"):
    try:
        if not os.path.exists(template_path):
            return None, f"Plantilla '{template_path}' no encontrada."

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 1. Datos Personales (Append to existing text)
        if 'nombre' in data_json and ws['D6'].value:
            ws['D6'] = f"{ws['D6'].value} {data_json['nombre']}"
            
        if 'cedula' in data_json and ws['D7'].value:
            ws['D7'] = f"{ws['D7'].value} {data_json['cedula']}"
            
        # 2. Idoneidad y Formación
        if 'idoneidad_texto' in data_json:
            ws['D10'] = data_json['idoneidad_texto']
            ws['D10'].alignment = Alignment(wrap_text=True, vertical='top')
            
        if 'formacion_texto' in data_json:
            ws['D13'] = data_json['formacion_texto']
            ws['D13'].alignment = Alignment(wrap_text=True, vertical='top')
            
        # 3. Tabla de Experiencia
        if 'experiencia_lista' in data_json:
            start_row = 21 # Nueva fila de inicio
            for i, exp in enumerate(data_json['experiencia_lista']):
                if i > 15: break # Limite de filas
                row = start_row + i
                
                # Validar fechas estrictas
                fecha_inicio = exp.get('fecha_inicio', '')
                fecha_fin = exp.get('fecha_fin', '')
                
                # Si falta alguna fecha completa, no escribir (aunque la IA ya debió filtrar)
                if len(fecha_inicio) == 10 and len(fecha_fin) == 10:
                    ws[f'D{row}'] = exp.get('empresa', '')
                    ws[f'E{row}'] = fecha_inicio
                    ws[f'F{row}'] = fecha_fin
                    ws[f'I{row}'] = exp.get('validada', 'NO') # Columna I para Validada
                
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), None
    except Exception as e:
        return None, str(e)

def clean_and_parse_json(text):
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            text = re.sub(r"```json\s*", "", text)
            text = re.sub(r"```\s*$", "", text)
            text = text.strip()
            text = text.replace('\n', '\\n').replace('\r', '').replace('\t', '\\t')
            return json.loads(text, strict=False)
        except:
            raise

# --- INTERFAZ PRINCIPAL ---
col1, col2 = st.columns([1, 1.2], gap="large")

with col1:
    st.subheader("📋 1. Perfil Requerido")
    tab_pdf, tab_text = st.tabs(["📄 Subir PDF", "✍️ Pegar Texto"])
    with tab_pdf:
        requisitos_pdf = st.file_uploader("Cargar PDF del Perfil", type=["pdf"], key="req_pdf")
    with tab_text:
        requisitos_text = st.text_area("Requisitos manuales", height=150, placeholder="Pegue aquí los requisitos...")

with col2:
    st.subheader("👤 2. Candidato y Soportes")
    c1, c2 = st.columns(2)
    with c1: nombre = st.text_input("Nombre Completo")
    with c2: identificacion = st.text_input("Identificación")
    
    st.info("📂 Soporta PDF, JPG y PNG")
    soportes = st.file_uploader("Cargar Evidencias (Diploma, Actas, Certificados)", 
                               type=["pdf", "jpg", "jpeg", "png"], 
                               accept_multiple_files=True, 
                               key="soportes")

# --- BOTÓN DE ACCIÓN ---
st.markdown("<br>", unsafe_allow_html=True)
if st.button("🚀 EVALUAR CANDIDATO", type="primary"):
    if not api_key:
        st.error("❌ Falta la API Key.")
    elif not (requisitos_pdf or requisitos_text):
        st.error("❌ Faltan los Requisitos.")
    elif not soportes:
        st.error("❌ Faltan los Soportes.")
    else:
        with st.spinner("🧠 Analizando documentos e imágenes... Calculando tiempos..."):
            try:
                # 1. Preparar Contexto de Requisitos
                req_content = ""
                if requisitos_pdf:
                    req_content += f"REQUISITOS (PDF): {extraer_texto_pdf(requisitos_pdf)}\n"
                if requisitos_text:
                    req_content += f"REQUISITOS (TXT): {requisitos_text}\n"

                # 2. Preparar Contenido Multimodal para Gemini
                gemini_content = []
                
                # Prompt del Sistema (Instrucciones)
                system_prompt = f"""
                Eres el Auditor de Contratación del SENA.
                
                OBJETIVO:
                Determinar si el candidato {nombre} (ID: {identificacion}) CUMPLE o NO CUMPLE con el perfil.
                
                REGLAS ESTRICTAS DE VALIDACIÓN:
                1. FECHAS EXACTAS: Solo acepta experiencias con fecha de inicio y fin completas (DD/MM/AAAA). 
                   - Si una certificación solo tiene MM/AAAA -> DESCARTARLA (No cuenta).
                   - Si no tiene fecha de fin (y no es actual) -> DESCARTARLA.
                2. TARJETA PROFESIONAL: Si el perfil exige Tarjeta Profesional, búscala en los soportes.
                   - Si la encuentras, extrae: "COPNIA [Número] [Fecha]".
                3. INSTRUCTORES: La experiencia como "Instructor SENA" o similar ES VÁLIDA como experiencia técnica.
                4. SUMATORIA: Suma solo los tiempos de certificaciones VÁLIDAS (con fechas completas).

                SALIDA JSON OBLIGATORIA:
                {{
                    "nombre": "{nombre}",
                    "cedula": "{identificacion}",
                    "concepto_final": "CUMPLE (Solo si cumple 100% formación Y tiempo total experiencia) o NO CUMPLE",
                    "idoneidad_texto": "CONCLUSIÓN: [CUMPLE/NO CUMPLE]. Justificación detallada. Si falta Tarjeta Profesional y se requiere, indicarlo.",
                    "formacion_texto": "Título Profesional + Fecha Grado. (Y Tarjeta Profesional si aplica).",
                    "experiencia_lista": [
                        {{
                            "empresa": "Nombre Empresa",
                            "fecha_inicio": "DD/MM/AAAA",
                            "fecha_fin": "DD/MM/AAAA",
                            "meses": 12,
                            "dias": 0,
                            "validada": "SI"
                        }}
                    ],
                    "analisis_detallado_markdown": "Tabla resumen en Markdown."
                }}
                """
                
                gemini_content.append(system_prompt)
                gemini_content.append(f"=== PERFIL REQUERIDO ===\n{req_content}")
                gemini_content.append("=== EVIDENCIAS DEL CANDIDATO ===")

                # Procesar Soportes (PDF Texto + Imágenes)
                for archivo in soportes:
                    if archivo.type == "application/pdf":
                        text = extraer_texto_pdf(archivo)
                        gemini_content.append(f"DOCUMENTO PDF ({archivo.name}):\n{text}")
                    elif archivo.type in ["image/png", "image/jpeg", "image/jpg"]:
                        img = cargar_imagen(archivo)
                        if img:
                            # Optimizar imagen antes de enviar
                            img_opt = optimize_image(img)
                            if img_opt:
                                gemini_content.append(f"IMAGEN ({archivo.name}):")
                                gemini_content.append(img_opt)

                # 3. Llamada al Modelo
                model = genai.GenerativeModel("gemini-2.0-flash", generation_config={"response_mime_type": "application/json"})
                response = model.generate_content(gemini_content)
                
                # 4. Procesar Respuesta
                data_json = clean_and_parse_json(response.text)
                
                # 5. Visualización
                st.markdown("<div class='result-container'>", unsafe_allow_html=True)
                
                concepto = data_json.get('concepto_final', 'NO CUMPLE').upper()
                color_banner = "#39A900" if "CUMPLE" in concepto and "NO" not in concepto else "#FC7323"
                icon_banner = "✅" if "CUMPLE" in concepto and "NO" not in concepto else "⚠️"
                
                st.markdown(f"""
                    <div style='background-color: {color_banner}; color: white; padding: 20px; border-radius: 12px; text-align: center; font-size: 24px; font-weight: 700; margin-bottom: 25px; box-shadow: 0 4px 12px rgba(0,0,0,0.15);'>
                        {icon_banner} {concepto}
                    </div>
                """, unsafe_allow_html=True)

                st.markdown("### 📊 Análisis de Idoneidad")
                if 'analisis_detallado_markdown' in data_json:
                    st.markdown(data_json['analisis_detallado_markdown'])
                else:
                    st.markdown(data_json.get('idoneidad_texto', ''))
                
                st.markdown("### 🗓️ Detalle de Experiencia (Sumatoria)")
                if 'experiencia_lista' in data_json and data_json['experiencia_lista']:
                    df_exp = pd.DataFrame(data_json['experiencia_lista'])
                    st.table(df_exp)
                else:
                    st.info("No se extrajo experiencia estructurada.")

                # Exportar
                excel_data, error_msg = fill_excel_template(data_json)
                if excel_data:
                    st.download_button(
                        label="📥 Descargar Concepto (Excel)",
                        data=excel_data,
                        file_name=f"IDONEIDAD_{nombre.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error(f"Error generando Excel: {error_msg}")
                
                st.markdown("</div>", unsafe_allow_html=True)

            except Exception as e:
                st.error(f"Ocurrió un error: {str(e)}")
                with st.expander("Ver detalle técnico"):
                    st.code(str(e))

